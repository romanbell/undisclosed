# TODO: Refactor type tagging methodology so that you can eventually query different pages by type (i.e. articles vs music vs...)

from pymongo import MongoClient
import pandas as pd
import json
from dotenv import load_dotenv
import os
import openpyxl

MONGO_DB_NAME = 'undisclosed_db'

# Cool sites
SITES_FILEPATH = 'cool_sites.csv'
MONGO_COOL_SITES_NAME = 'cool_sites'


# Goodreads
GOODREADS_FILEPATH = 'Goodreads_Export_20221107.xlsx'
MONGO_GOODREADS_NAME = 'goodreads'

# Visual Artists
VIZ_FILEPATH = 'visual_artists.csv'
MONGO_VIZ_NAME = 'visual_artists'

# Spotify Wrapped
WRAPPED_FILEPATHS = {
    2020: '2020_wrapped.json',
    2021: '2021_wrapped.json',
    2022: '2022_wrapped.json'
}
WRAPPED_NAME = 'wrapped'

class MongoLoader():
    def __init__(self):
        print(os.getcwd())
        load_dotenv(os.path.join(os.getcwd(), '.env'))
        self.MONGODB_PASSWORD = os.environ.get('MONGO_PWD')
        if self.MONGODB_PASSWORD is None:
            raise Exception('Error loading env vars')

    def trigger_cool_sites_load(self):
        print('Loading all cool sites records into MongoDB')
        cool_sites_df = self.get_cool_sites_to_json_df()
        mongo_conn = self.connect_to_db()
        self.delet_all_existing_records(conn=mongo_conn, collection_to_delete=MONGO_COOL_SITES_NAME)
        self.upload_json_df_to_db(conn=mongo_conn, df=cool_sites_df, collection_name=MONGO_COOL_SITES_NAME)
        print('Cool sites records loaded successfully')

    def trigger_goodreads_load(self):
        print('Loading all goodreads records into MongoDB')
        goodreads_df = self.get_goodreads_to_json_df()
        mongo_conn = self.connect_to_db()
        self.delet_all_existing_records(conn=mongo_conn, collection_to_delete=MONGO_GOODREADS_NAME)
        self.upload_json_df_to_db(conn=mongo_conn, df=goodreads_df, collection_name=MONGO_GOODREADS_NAME)
        print('Cool sites records loaded successfully')

    def trigger_vartists_load(self):
        print('Loading all visual artists into MongoDB')
        vartists_df = self.get_visual_artists_to_json_df()
        mongo_conn = self.connect_to_db()
        self.delet_all_existing_records(conn=mongo_conn, collection_to_delete=MONGO_VIZ_NAME)
        self.upload_json_df_to_db(conn=mongo_conn, df=vartists_df, collection_name=MONGO_VIZ_NAME)
        print('Visual artists records loaded successfully')

    def trigger_wrapped_load(self):
        print('Loading all spotify wrapped data into MongoDB')
        mongo_conn = self.connect_to_db()
        self.delet_all_existing_records(conn=mongo_conn, collection_to_delete=WRAPPED_NAME)
        self.parse_and_upload_wrapped_json(conn=mongo_conn)
        print("Wrapped loaded successfully")

    def parse_and_upload_wrapped_json(self, conn):
        db = conn.get_database(MONGO_DB_NAME)
        spotify_collection = db[WRAPPED_NAME]
        for yr, filepath in WRAPPED_FILEPATHS.items():
            with open(os.path.join('loaders', filepath), encoding='UTF-8') as d:
                spotify_json = json.load(d)
            spotify_tracks = spotify_json.get('items')
            rank = 1
            for song in spotify_tracks:
                song['rank'] = rank
                song['year'] = yr
                spotify_collection.insert_one(song)
                rank += 1
            print(f"All wrapped tracks for {yr} inserted successfully")

    def get_visual_artists_to_json_df(self):
        def _parse_location_list(l):
            if len(l) <= 1:
                return ''
            else:
                return l.split(';')

        vart_df = pd.read_csv(os.path.join(os.getcwd(), 'loaders', VIZ_FILEPATH)).fillna('')
        vart_df.loc[:, 'loc_list_parsed'] = vart_df['loc_list'].apply(lambda x: _parse_location_list(x))
        vart_df.loc[:, 'date_added'] = pd.to_datetime(vart_df['date_added']) 
        vart_df.loc[:, 'date_added_str'] = vart_df['date_added'].dt.strftime('%Y%m%d')
        vart_df = vart_df.drop('loc_list', axis=1)
        vart_df['json'] = vart_df.apply(lambda x: x.to_json(), axis=1)
        return vart_df

    def get_cool_sites_to_json_df(self):

        def _parse_comma_list(l):
            if type(l) != str or l == None or l == '' or l == ' ':
                return []
            else:
                return l.split(",")

        sites_df = pd.read_csv(os.path.join(os.getcwd(), 'loaders', SITES_FILEPATH))
        sites_df.loc[:, 'date_added'] = pd.to_datetime(sites_df['date_added']) 
        sites_df.loc[:, 'date_added_str'] = sites_df['date_added'].dt.strftime('%Y%m%d')
        sites_df.loc[:, 'creator_name'] = sites_df['creator_name'].apply(lambda x: _parse_comma_list(x))
        sites_df.loc[:, 'creator_link'] = sites_df['creator_link'].apply(lambda x: _parse_comma_list(x))
        sites_df = sites_df.fillna('').rename(columns={'loc_a': 'city', 'loc_b': 'country'})
        sites_df['json'] = sites_df.apply(lambda x: x.to_json(), axis=1)
        return sites_df

    def get_goodreads_to_json_df(self):
        wb = openpyxl.load_workbook(os.path.join(os.getcwd(), 'loaders', GOODREADS_FILEPATH))
        ws = wb.active

        book_hypertext, author_hypertext = [],  []
        for r in range(2, ws.max_row):
        #     print(r, book_hypertext, author_hypertext)
            if ws.cell(row=r, column=1).hyperlink == None:
                book_hypertext.append('')
                author_hypertext.append('')
            else:
                try:
                    book_hypertext.append(str(ws.cell(row=r, column=1).hyperlink.target).replace('\\', ''))
                    author_hypertext.append(str(ws.cell(row=r, column=2).hyperlink.display).replace('\\', ''))
                except Exception as e:
                    book_hypertext.append(str(ws.cell(row=r, column=1).hyperlink)+'')
                    author_hypertext.append(str(ws.cell(row=r, column=2).hyperlink)+'')
                    
        gr_df = pd.read_excel(os.path.join(os.getcwd(), 'loaders', GOODREADS_FILEPATH))

        gr_df.loc[:, 'Title Link'] = book_hypertext + ['']
        gr_df.loc[:, 'Author Link'] = author_hypertext + ['']
        gr_df = gr_df.dropna(subset=['Title'], axis=0).reset_index(drop=True)

        def _remove_edit_string(s):
            return s[:-7]

        def _remove_asterisk(s):
            return s.replace('*', '')

        def _extract_id(s):
            return s[36:]

        gr_df.loc[:, 'Date Read'] = gr_df['Date Read'].apply(lambda x: _remove_edit_string(x))
        gr_df.loc[:, 'Author'] = gr_df['Author'].apply(lambda x: _remove_asterisk(x))
        gr_df.loc[:, 'id'] = gr_df['Title Link'].apply(lambda x: _extract_id(x))
        gr_df['Date Read'] = gr_df['Date Read'].replace('not set', None)
        gr_df['Date Read'] = pd.to_datetime(gr_df['Date Read'], format="%b %d, %Y")
        gr_df.loc[:, 'Date Read String'] = gr_df['Date Read'].dt.strftime('%Y%m%d').replace('20201223', '00000000').fillna('00000000')
        gr_df.loc[:, 'Date Read Rank'] = gr_df.index.tolist()
        gr_df = gr_df.rename(columns={'Title': 'title', 'Author': 'author', 'Date Read': 'date_read', 'Title Link': 'title_link',
                                     'Author Link': 'author_link', 'Date Read Rank': 'date_read_rank', 'Date Read String': 'date_read_string'})
        gr_df['json'] = gr_df.apply(lambda x: x.to_json(), axis=1)
        return gr_df

    def connect_to_db(self):
        try:
            conn = MongoClient(f"mongodb+srv://romanbell:{self.MONGODB_PASSWORD}@cluster0.qzaqz.mongodb.net/test")
            print("MongoDB connection successfully established...")
            return conn
        except:  
            print("Could not connect to MongoDB")

    def delet_all_existing_records(self, conn, collection_to_delete):
        db = conn.get_database(MONGO_DB_NAME)
        collection = db.get_collection(collection_to_delete)
        x = collection.delete_many({})
        print(f"Successfully deleted {x.deleted_count} documents from {collection_to_delete}")

    def upload_json_df_to_db(self, conn, df, collection_name, json_col_name='json'):

        db = conn.get_database(MONGO_DB_NAME)
        collection = db.get_collection(collection_name)

        for idx, record in enumerate(df[json_col_name]):
            json_record = json.loads(record)
            collection.insert_one(json_record)   
            print(f'record {idx} inserted successfully to {collection_name} collection')   


if __name__ == '__main__':
    its_mongo_time = MongoLoader()
    # its_mongo_time.trigger_cool_sites_load()
    # its_mongo_time.trigger_goodreads_load()
    # its_mongo_time.trigger_vartists_load()
    its_mongo_time.trigger_wrapped_load()
    print('load complete')